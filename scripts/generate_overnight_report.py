#!/usr/bin/env python3
"""
Generate an overnight market report workbook similar to the sample file.

Data sources:
- https://www.marketindex.com.au/commodities
- https://www.investing.com

Usage:
    python3 generate_overnight_report.py
    python3 generate_overnight_report.py --output overnight_data.xlsx
    python3 generate_overnight_report.py --prefer-browser
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import requests
from bs4 import BeautifulSoup
from bs4 import Comment
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
import yfinance as yf


USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/135.0.0.0 Safari/537.36"
)

REQUEST_HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Language": "en-AU,en-US;q=0.9,en;q=0.8",
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "Referer": "https://www.google.com/",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}

MARKETINDEX_COMMODITIES_URL = "https://www.marketindex.com.au/commodities"
TRADING_ECONOMICS_IRON_ORE_URL = "https://tradingeconomics.com/commodity/iron-ore"


@dataclass(frozen=True)
class Quote:
    last: float | None
    change: float | None
    pct_change: float | None


@dataclass(frozen=True)
class InvestingPage:
    url: str
    value_transform: Callable[[Quote], Quote] | None = None


@dataclass(frozen=True)
class YFinanceSource:
    symbol: str
    value_transform: Callable[[Quote], Quote] | None = None


def normalize_number(raw: str | None) -> float | None:
    if raw is None:
        return None
    cleaned = (
        raw.strip()
        .replace(",", "")
        .replace("%", "")
        .replace("−", "-")
        .replace("+", "")
        .replace("¥", "")
        .replace("$", "")
    )
    if not cleaned:
        return None
    return float(cleaned)


def quote_from_strings(last: str | None, change: str | None, pct: str | None) -> Quote:
    last_value = normalize_number(last)
    change_value = normalize_number(change)
    pct_value = normalize_number(pct)
    if pct_value is not None:
        pct_value /= 100
    return Quote(last=last_value, change=change_value, pct_change=pct_value)


def invert_quote(quote: Quote) -> Quote:
    if quote.last in (None, 0):
        return Quote(last=None, change=None, pct_change=None)
    prev = None if quote.change is None else quote.last - quote.change
    inverted_last = 1 / quote.last
    if prev in (None, 0):
        return Quote(last=inverted_last, change=None, pct_change=None)
    inverted_prev = 1 / prev
    inverted_change = inverted_last - inverted_prev
    inverted_pct = None if inverted_prev == 0 else inverted_change / inverted_prev
    return Quote(last=inverted_last, change=inverted_change, pct_change=inverted_pct)


def yield_quote_to_decimal(quote: Quote) -> Quote:
    return Quote(
        last=None if quote.last is None else quote.last / 100,
        change=None if quote.change is None else quote.change / 100,
        pct_change=quote.pct_change,
    )


def fetch_html_via_browser(url: str, timeout: int) -> str:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError(
            "Playwright is required for browser fallback. "
            "Install it with `pip install playwright` and run "
            "`python -m playwright install chromium`."
        ) from exc

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page(user_agent=USER_AGENT, locale="en-AU")
        page.goto(url, wait_until="domcontentloaded", timeout=timeout * 1000)
        page.wait_for_timeout(2500)
        html = page.content()
        browser.close()
    return html


def fetch_html(session: requests.Session, url: str, timeout: int, prefer_browser: bool) -> str:
    if prefer_browser:
        return fetch_html_via_browser(url, timeout)

    response = session.get(url, timeout=timeout, headers=REQUEST_HEADERS)
    if response.status_code == 403:
        return fetch_html_via_browser(url, timeout)
    response.raise_for_status()
    return response.text


def flatten_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    for comment in soup.find_all(string=lambda text: isinstance(text, Comment)):
        comment.extract()
    return re.sub(r"\s+", " ", soup.get_text(separator=" ", strip=True))


def parse_investing_quote(html: str, url: str) -> Quote:
    price_match = re.search(
        r'text-black font-bold">\s*([+\-−]?\d[\d,\.]*)\s*</div>',
        html,
    )
    change_match = re.search(
        r'<div class="font-medium mr-1">\s*([+\-−]?\d[\d,\.]*)\s*</div>',
        html,
    )
    pct_match = re.search(
        r'<span>\s*([+\-−]?\d[\d,\.]*%)\s*</span>',
        html,
    )
    if price_match:
        return quote_from_strings(
            price_match.group(1),
            change_match.group(1) if change_match else None,
            pct_match.group(1) if pct_match else None,
        )

    text = flatten_text(html)
    label = url.rsplit("/", 1)[-1].replace("-", " ").title() if url.startswith("http") else url
    patterns = [
        rf"{re.escape(label)}\s+([+\-−]?\d[\d,\.]*)"
        rf"(?:\s+ASX .*? Companies)?"
        rf"(?:\s+([+\-−]?\d[\d,\.]*)\s+\(([+\-−]?\d[\d,\.]*%)\))?",
        r"Add to Watchlist\s+([+\-−]?\d[\d,\.]*)\s+([+\-−]?\d[\d,\.]*)\s*\(?([+\-−]?\d[\d,\.]*%)\)?",
        r"Add to/Remove from Watchlist Add to Watchlist\s+([+\-−]?\d[\d,\.]*)\s+([+\-−]?\d[\d,\.]*)\s+([+\-−]?\d[\d,\.]*%)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return quote_from_strings(match.group(1), match.group(2), match.group(3))
    raise ValueError(f"Could not parse Investing quote from {url}")


def parse_market_index_commodities(html: str) -> dict[str, Quote]:
    soup = BeautifulSoup(html, "html.parser")
    commodity_names = [
        "Iron Ore",
        "Crude Oil",
        "Gold",
        "Copper",
        "Aluminium",
        "Nickel",
        "Zinc",
        "Lead",
    ]
    parsed: dict[str, Quote] = {}
    for name in commodity_names:
        title = soup.find(
            "div",
            string=lambda value, commodity=name: value and value.strip() == commodity,
        )
        if title is None:
            continue
        card = title.find_parent(
            "div",
            class_=lambda classes: classes
            and "rounded-lg" in classes
            and "bg-white" in classes
            and "border" in classes,
        )
        if card is None:
            continue
        parsed[name] = parse_investing_quote(str(card), name)

    return parsed


def quote_from_previous_close(last: float | None, previous_close: float | None) -> Quote:
    if last is None:
        return Quote(last=None, change=None, pct_change=None)
    if previous_close in (None, 0):
        return Quote(last=last, change=None, pct_change=None)
    change = last - previous_close
    pct_change = change / previous_close
    return Quote(last=last, change=change, pct_change=pct_change)


def fetch_yfinance_quote(symbol: str, value_transform: Callable[[Quote], Quote] | None = None) -> Quote:
    history = yf.Ticker(symbol).history(period="5d", interval="1d", auto_adjust=False)
    if history.empty:
        return Quote(last=None, change=None, pct_change=None)

    closes = history["Close"].dropna()
    if closes.empty:
        return Quote(last=None, change=None, pct_change=None)

    last = float(closes.iloc[-1])
    previous_close = float(closes.iloc[-2]) if len(closes) > 1 else None
    quote = quote_from_previous_close(last, previous_close)
    if value_transform is not None:
        quote = value_transform(quote)
    return quote


def fetch_tradingeconomics_iron_ore(session: requests.Session, timeout: int) -> Quote:
    response = session.get(
        TRADING_ECONOMICS_IRON_ORE_URL,
        timeout=timeout,
        headers=REQUEST_HEADERS,
    )
    response.raise_for_status()
    text = flatten_text(response.text)

    # summary_match = re.search(
    #     r"Iron Ore rose to ([+\-−]?\d[\d,\.]*) .*? up ([+\-−]?\d[\d,\.]*%) from the previous day",
    #     text,
    # )
    # if summary_match:
    #     last = summary_match.group(1)
    #     pct = summary_match.group(2)
    #     pct_value = normalize_number(pct)
    #     last_value = normalize_number(last)
    #     change_value = None
    #     if last_value is not None and pct_value is not None:
    #         previous_close = last_value / (1 + (pct_value / 100))
    #         change_value = last_value - previous_close
    #     return Quote(
    #         last=last_value,
    #         change=change_value,
    #         pct_change=None if pct_value is None else pct_value / 100,
    #     )

    row_match = re.search(
        r"Iron Ore\s+([+\-−]?\d[\d,\.]*)\s+([+\-−]?\d[\d,\.]*)\s+([+\-−]?\d[\d,\.]*%)",
        text,
    )
    if row_match:
        return quote_from_strings(row_match.group(1), row_match.group(2), row_match.group(3))

    raise ValueError("Could not parse Iron Ore quote from Trading Economics")


INVESTING_SOURCES: dict[str, InvestingPage] = {
    "xjo": InvestingPage("https://www.investing.com/indices/aus-200"),
    "spi": InvestingPage("https://www.investing.com/indices/australia-200-futures"),
    "vix": InvestingPage("https://www.investing.com/indices/volatility-s-p-500"),
    "asx200vix": InvestingPage("https://www.investing.com/indices/s-p-asx-200-vix"),
    "au10y": InvestingPage(
        "https://www.investing.com/rates-bonds/australia-10-year-bond-yield",
        value_transform=yield_quote_to_decimal,
    ),
    "dow": InvestingPage("https://www.investing.com/indices/us-30"),
    "sp500": InvestingPage("https://www.investing.com/indices/us-spx-500"),
    "nasdaq": InvestingPage("https://www.investing.com/indices/nasdaq-composite"),
    "us10y": InvestingPage(
        "https://www.investing.com/rates-bonds/u.s.-10-year-bond-yield",
        value_transform=yield_quote_to_decimal,
    ),
    "ftse100": InvestingPage("https://www.investing.com/indices/uk-100"),
    # "stoxx600": InvestingPage("https://www.investing.com/indices/stoxx-600"),
    "dax": InvestingPage("https://www.investing.com/indices/germany-30"),
    "cac40": InvestingPage("https://www.investing.com/indices/france-40"),
    "nikkei": InvestingPage("https://www.investing.com/indices/japan-ni225"),
    "hangseng": InvestingPage("https://www.investing.com/indices/hang-sen-40"),
    "audusd": InvestingPage("https://www.investing.com/currencies/aud-usd"),
    "eurusd": InvestingPage("https://www.investing.com/currencies/eur-usd"),
    "audgbp": InvestingPage(
        "https://www.investing.com/currencies/gbp-aud",
        value_transform=invert_quote,
    ),
}

YFINANCE_SOURCES: dict[str, YFinanceSource] = {
    "stoxx600": YFinanceSource("^STOXX"),
}

MARKET_INDEX_KEY_MAP: dict[str, str] = {
    "mi_crude_oil": "Crude Oil",
    "mi_gold": "Gold",
    # "mi_iron_ore": "Iron Ore",
    "mi_copper": "Copper",
    "mi_aluminium": "Aluminium",
    "mi_nickel": "Nickel",
    "mi_zinc": "Zinc",
    "mi_lead": "Lead",
}


ROW_DATA = [
    ("A1", "Overnight Data 隔夜数据", None),
    ("A3", "S&P/ASX 200", None),
    ("B3", "Close 收盘价", None),
    ("C3", "Change 涨跌", None),
    ("D3", "Pct% 涨跌幅", None),
    ("A4", "XJO 澳洲200指数", "xjo"),
    ("A5", "SPI 澳洲200指数期货", "spi"),
    ("A6", "VIX Index 恐慌指数", "asx200vix"),
    ("A7", "AUS 10 Yr Bonds\n澳洲10年国债收益率", "au10y"),
    ("A8", "US Market 美国市场", None),
    ("A9", "Dow Jones 道琼斯指数", "dow"),
    ("A10", "S&P 500 标普500指数", "sp500"),
    ("A11", "Nasdaq 纳斯达克指数", "nasdaq"),
    ("A12", "U.S. 10 Yr Bonds\n美国10年国债收益率", "us10y"),
    ("A13", "VIX Index 恐慌指数", "vix"),
    ("A14", "Regional Indices 全球主要股指", None),
    ("A15", "FTSE 100 \n英国富时100指数", "ftse100"),
    ("A16", "STOXX 600 欧洲600指数", "stoxx600"),
    ("A17", "DAX Index 德国DAX指数", "dax"),
    ("A18", "CAC 40 法国CAC40指数", "cac40"),
    ("A19", "Nikkei 日本日经225指数", "nikkei"),
    ("A20", "Hang Seng 香港恒生指数", "hangseng"),
    ("A22", "Commidities 大宗商品", None),
    ("B22", "Last 最新价", None),
    ("C22", "Change 涨跌", None),
    ("D22", "Pct% 涨跌幅", None),
    ("A23", "Brent Crude - US$/bbl \n布伦特原油", "mi_crude_oil"),
    ("A24", "Gold - US$/oz 黄金", "mi_gold"),
    ("A26", "Iron Ore(US$/t) 铁矿石", "mi_iron_ore"),
    ("A27", "Copper(US$/lb) 铜", "mi_copper"),
    ("A28", "Aluminium(US$/t) 铝", "mi_aluminium"),
    ("A29", "Nickel(US$/t) 镍", "mi_nickel"),
    ("A30", "Zinc(US$/t) 锌", "mi_zinc"),
    ("A31", "Lead(US$/t) 铅", "mi_lead"),
    ("A33", "Currencies 汇率", None),
    ("A34", "AUD/USD 澳元/美元", "audusd"),
    ("A35", "EUR/USD 欧元/美元", "eurusd"),
    ("A36", "AUD/GBP 澳元/英镑", "audgbp"),
]


def collect_quotes(timeout: int, prefer_browser: bool) -> dict[str, Quote]:
    session = requests.Session()
    quotes: dict[str, Quote] = {}

    market_index_html = fetch_html(
        session,
        MARKETINDEX_COMMODITIES_URL,
        timeout=timeout,
        prefer_browser=prefer_browser,
    )
    market_index_quotes = parse_market_index_commodities(market_index_html)
    for output_key, commodity_name in MARKET_INDEX_KEY_MAP.items():
        quotes[output_key] = market_index_quotes.get(
            commodity_name,
            Quote(last=None, change=None, pct_change=None),
        )

    for key, source in YFINANCE_SOURCES.items():
        quotes[key] = fetch_yfinance_quote(
            source.symbol,
            value_transform=source.value_transform,
        )

    quotes["mi_iron_ore"] = fetch_tradingeconomics_iron_ore(session, timeout)

    for key, page in INVESTING_SOURCES.items():
        html = fetch_html(
            session,
            page.url,
            timeout=timeout,
            prefer_browser=prefer_browser,
        )
        quote = parse_investing_quote(html, page.url)
        if page.value_transform is not None:
            quote = page.value_transform(quote)
        quotes[key] = quote

    return quotes


def apply_layout(ws) -> None:
    ws.title = "Sheet1"
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A8:D8")
    ws.merge_cells("A14:D14")

    ws.column_dimensions["A"].width = 24.16
    ws.column_dimensions["B"].width = 12.83
    ws.column_dimensions["C"].width = 12.83
    ws.column_dimensions["D"].width = 12.50

    ws.row_dimensions[1].height = 23
    ws.row_dimensions[2].height = 23
    ws.row_dimensions[3].height = 17
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 17
    ws.row_dimensions[12].height = 34
    ws.row_dimensions[14].height = 17
    ws.row_dimensions[15].height = 34
    ws.row_dimensions[22].height = 17
    ws.row_dimensions[23].height = 34
    ws.row_dimensions[33].height = 17


def style_sheet(ws) -> None:
    title_font = Font(name="Arial", size=18)
    body_font = Font(name="Arial", size=12)
    header_font = Font(name="Arial", size=12, bold=True, color="4F81BD")
    positive_font = Font(name="Arial", size=12, color="008000")
    negative_font = Font(name="Arial", size=12, color="C00000")
    medium_bottom = Border(bottom=Side(style="medium"))

    for row in range(1, 37):
        for col in "ABCD":
            ws[f"{col}{row}"].font = body_font

    for row in (3, 22):
        for cell in ws[row]:
            cell.font = header_font
            cell.border = medium_bottom

    for row in (8, 14, 33):
        for cell in ws[row]:
            cell.font = header_font
            cell.border = medium_bottom
        ws[f"A{row}"].alignment = Alignment(horizontal="left")

    ws["A1"].font = title_font

    for row in range(4, 37):
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    for row in range(4, 37):
        value_cell = ws[f"B{row}"]
        change_cell = ws[f"C{row}"]
        pct_cell = ws[f"D{row}"]

        label = ws[f"A{row}"].value or ""
        is_bond_row = "Bonds" in label

        if is_bond_row:
            value_cell.number_format = "0.000%"
            change_cell.number_format = "0.000%"
        else:
            value_cell.number_format = "#,##0.00"
            change_cell.number_format = "#,##0.00"

        pct_cell.number_format = "0.00%"

        if change_cell.value is not None and isinstance(change_cell.value, (int, float)):
            if change_cell.value > 0:
                change_cell.font = positive_font
            elif change_cell.value < 0:
                change_cell.font = negative_font

        if pct_cell.value is not None and isinstance(pct_cell.value, (int, float)):
            if pct_cell.value > 0:
                pct_cell.font = positive_font
            elif pct_cell.value < 0:
                pct_cell.font = negative_font


def populate_sheet(ws, quotes: dict[str, Quote]) -> None:
    for cell_ref, label, key in ROW_DATA:
        ws[cell_ref] = label
        if not key:
            continue

        row = ws[cell_ref].row
        quote = quotes.get(key, Quote(last=None, change=None, pct_change=None))
        ws[f"B{row}"] = quote.last
        ws[f"C{row}"] = quote.change
        ws[f"D{row}"] = quote.pct_change


def build_workbook(quotes: dict[str, Quote]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    apply_layout(sheet)
    populate_sheet(sheet, quotes)
    style_sheet(sheet)
    return workbook


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate the overnight market report workbook.")
    parser.add_argument(
        "--output",
        default="overnight_data_generated.xlsx",
        help="Output workbook path",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=20,
        help="HTTP request timeout in seconds",
    )
    parser.add_argument(
        "--prefer-browser",
        action="store_true",
        help="Use Playwright/Chromium for all requests instead of plain HTTP.",
    )
    args = parser.parse_args()

    quotes = collect_quotes(timeout=args.timeout, prefer_browser=args.prefer_browser)
    workbook = build_workbook(quotes)
    output_path = Path(args.output).expanduser().resolve()
    workbook.save(output_path)
    print(f"Saved workbook to {output_path}")


if __name__ == "__main__":
    main()
